import datetime
from openpyxl import load_workbook
from openpyxl import Workbook

# 读取文件以及相应的表
wb = load_workbook('courses.xlsx')
students_sheet = wb['students']
time_sheet = wb['time']


def combine():
    # 创建 combine 表
    combine_sheet = wb.create_sheet(title='combine')
    # 向 combine 表添加表头
    combine_sheet.append(['创建时间', '课程名称', '学习人数', '学习时间'])
    # 合并表格并依次添加到 combine 表中
    for stu in students_sheet.values:
        # 去掉包含表头的一行
        if stu[2] != '学习人数':
            # 遍历匹配学习时间数据
            for time in time_sheet.values:
                if time[1] == stu[1]:
                    # 添加记录到 combine 表中
                    combine_sheet.append(list(stu) + [time[2]])
    # 覆盖保存 courses.xlsx
    wb.save('courses.xlsx')


def split():
    combine_sheet = wb['combine']
    # 存储 combine 表中的年份
    split_name = []
    # 遍历表获取每条数据对应的年份
    for item in combine_sheet.values:
        if item[0] != '创建时间':
            split_name.append(item[0].strftime("%Y"))

    # 分别储存数据
    for name in set(split_name):
        # 创建文件
        wb_temp = Workbook()
        # 删除已有的默认 Sheet 表
        wb_temp.remove(wb_temp.active)
        # 创建相应年份命名的表
        ws = wb_temp.create_sheet(title=name)
        # 写入相应年份的数据
        for item_by_year in combine_sheet.values:
            if item_by_year[0] != '创建时间':
                if item_by_year[0].strftime("%Y") == name:
                    ws.append(item_by_year)
        # 存储相应年份的数据文件
        wb_temp.save('{}.xlsx'.fromat(name))

# 执行
if __name__ == '__main__':
    combine()
    split()
